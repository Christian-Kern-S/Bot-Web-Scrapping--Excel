import requests
from bs4 import BeautifulSoup
import openpyxl


# abre o arquivo xlsx
workbook = openpyxl.load_workbook('custos.xlsx')

# seleciona a planilha ativa
sheet = workbook.active


def get_price(link):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 OPR/96.0.0.0"

    }
    response = requests.get(link, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
    print(soup)
    price_element = soup.find('span', class_='price')
    print(price_element)
    if price_element is not None:
        price_text = price_element.text.strip()
        price_num = float(price_text.split(
            ' ')[-1].replace('.', '').replace('R$', '').replace(',', '.'))
        return price_num
    else:
        return 'Preço não encontrado'


# percorre as linhas da coluna A, a partir da segunda linha
for row in sheet.iter_rows(min_row=2, min_col=1):
    # obtém o link da célula atual
    link = row[0].value
    if link is None:
        break

    # faz a requisição ao site do AliExpress
    if link is not None:
        price = get_price(link)
    else:
        price = 'Preço não encontrado'

    # insere o preço na coluna C, na mesma linha do link
    sheet.cell(row=row[0].row, column=3).value = price

    # obtém os valores das colunas B e C, na mesma linha
    value_b = row[1].value
    value_c = row[2].value

    # calcula a diferença entre os valores
    try:
        diff = float(value_c) - float(value_b)
    except ValueError:
        diff = 0.0

    # insere a diferença na coluna D, na mesma linha do link
    sheet.cell(row=row[0].row, column=4).value = diff

    # verifica se houve aumento ou diminuição
    if diff > 0:
        status = 'Aumentou'
        font_color = '000000'  # preto
        fill_color = 'FF0000'  # vermelho
    elif diff < 0:
        status = 'Diminuiu'
        font_color = '000000'  # preto
        fill_color = '008000'  # verde
    else:
        status = ''
        font_color = '000000'  # preto
        fill_color = 'FFFFFF'  # branco

    # insere a informação na coluna E, na mesma linha do link
    sheet.cell(row=row[0].row, column=5).value = status
    sheet.cell(row=row[0].row, column=5).font = openpyxl.styles.Font(
        color=font_color)
    sheet.cell(row=row[0].row, column=5).fill = openpyxl.styles.PatternFill(
        start_color=fill_color, end_color=fill_color, fill_type='solid')

# salva as alterações no arquivo xlsx
workbook.save('custos.xlsx')
